import openpyxl
import hashlib
import re

# ==========================================
# Configuration
# ==========================================
ASSESSMENT_TYPE = "Assignment-01"
SEMESTER_NAME = "Summer 2025"
COURSE_CODE = "MAT215"
COURSE_NAME = r"Complex Variables \& Laplace Transform"
SECTION = "12"

TEMPLATE_PATH = "question_template.txt"
OUTPUT_PATH = f"{ASSESSMENT_TYPE}.tex"
WORKBOOK_PATH = "course-attendee.xlsx"
SHEET_NAME = "Worksheet"


# ==========================================
# Load Workbook & Sheet
# ==========================================
workbook = openpyxl.load_workbook(WORKBOOK_PATH)
sheet = workbook[SHEET_NAME]


# ==========================================
# Utility Functions
# ==========================================
def generate_integers_range(ID, variation, n, a, b):
    """
    Generate deterministic pseudo-random integers from ID and variation.
    """
    input_string = f"{ID}{variation}"
    hash_value = hashlib.md5(input_string.encode()).hexdigest()
    return [(int(hash_value[i:i + 2], 32) % (b - a + 1)) + a for i in range(0, n * 2, 2)]


def replace_template_placeholders(template_text, **kwargs):
    """
    Replace placeholders in the LaTeX template with provided keyword arguments.
    """
    for key, value in kwargs.items():
        template_text = template_text.replace(f"@{key}@", str(value))
    return template_text


def append_to_output_file(content):
    """
    Append text content to the LaTeX output file.
    """
    with open(OUTPUT_PATH, 'a', encoding='utf-8') as f_out:
        f_out.write(content)


# ==========================================
# Prepare LaTeX Template
# ==========================================
with open(TEMPLATE_PATH, 'r', encoding='utf-8') as f_template:
    template = f_template.read().replace('\r\n', '\n')

header_match = re.search(r"^(.*?\\begin\{document\})", template, re.DOTALL)
body_match = re.search(r"\\begin\{document\}(.*?)\\end\{document\}", template, re.DOTALL)

if not (header_match and body_match):
    raise ValueError("Template must contain a valid LaTeX document structure with \\begin{document} and \\end{document}.")

latex_header = header_match.group(1).strip() + "\n\n"
latex_body = body_match.group(1).strip()
latex_footer = "\n\n\\end{document}"


# ==========================================
# Write Header
# ==========================================
with open(OUTPUT_PATH, "w", encoding="utf-8") as f_out:
    f_out.write(latex_header)


# ==========================================
# Main Loop
# ==========================================
row = 2
count = 0

while True:
    ID = sheet.cell(row=row, column=1).value
    if ID is None:
        print(f"\n✅ Total {count} questions generated.")
        break

    Name = sheet.cell(row=row, column=2).value or "Unknown"
    print(f"Processing: {Name}")

    replaced_text = replace_template_placeholders(latex_body,
        #Placeholder = Value
        Name=Name,
        ID=ID,
        Section=SECTION,
        Course_Name=COURSE_NAME,
        Course_Code=COURSE_CODE,
        Semester_Name=SEMESTER_NAME,
        Assesment_Type=ASSESSMENT_TYPE
    )

    append_to_output_file(replaced_text + "\n\n")
    count += 1
    row += 1


# ==========================================
# Write Footer
# ==========================================
append_to_output_file(latex_footer)

print(f"🎉 LaTeX file '{OUTPUT_PATH}' generated successfully.")
