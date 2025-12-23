#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Author: Partho Sutra Dhor
Email: parthosutradhor@gmail.com, partho.dhor@bracu.ac.bd
Created: October 2025
Description: Automated LaTeX question booklet generator
"""

import glob
import hashlib
import os
import re
import shutil
import subprocess
import time
from datetime import datetime

import openpyxl
from sympy import I, Rational, cos, latex, pi, simplify, sin

# ============================================================
# CONFIGURATION
# ============================================================
ASSESSMENT_TYPE = "Assignment - 02"
SEMESTER_NAME = "Fall 2025"
COURSE_CODE = "MAT 215"
COURSE_NAME = r"Complex Variables and Laplace Transformations"
SECTION = "12"
TOTAL_POINTS = "100"

TEMPLATE_PATH = "question_template - 02.tex"
WORKBOOK_PATH = "course-attendee.xlsx"
SHEET_NAME = "Worksheet"
LOGO_FILE = "Brac_University_Logo.png"

START_ROW = 2  # Data starts here (1-indexed, header on row 1)


# ============================================================
# GENERIC UTILITIES
# ============================================================
def generate_integers_range(ID, variation, n, a, b):
    """
    Deterministic pseudo-random integers based on (ID, variation).
    Uses MD5 -> hex pairs -> base-32 -> mod range mapping.
    """
    seed = f"{ID}{variation}"
    h = hashlib.md5(seed.encode()).hexdigest()
    return [
        (int(h[i:i + 2], 32) % (b - a + 1)) + a
        for i in range(0, n * 2, 2)
    ]

def generate_integer(ID, variation, a, b):
    """
    Deterministic pseudo-random integers based on (ID, variation).
    Uses MD5 -> hex pairs -> base-32 -> mod range mapping.
    """
    seed = f"{ID}{variation}"
    h = hashlib.md5(seed.encode()).hexdigest()
    return (int(h[0:2], 32) % (b - a + 1)) + a


def replace_placeholders(template_text, **kwargs):
    """Replace @Key@ placeholders in a LaTeX template."""
    for key, val in kwargs.items():
        template_text = template_text.replace(f"@{key}@", str(val))
    return template_text


def compile_latex_to_pdf(tex_path):
    """
    Compile .tex -> .pdf using pdflatex twice.
    Prints a short log tail if compilation fails.
    """
    tex_dir = os.path.dirname(tex_path) or "."
    tex_file = os.path.basename(tex_path)
    pdf_path = os.path.splitext(tex_path)[0] + ".pdf"

    cmd = ["pdflatex", "-interaction=nonstopmode", tex_file]

    build_ok = False
    last_stdout = ""
    for _ in range(2):
        result = subprocess.run(
            cmd, cwd=tex_dir, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, text=True
        )
        last_stdout = result.stdout
        if result.returncode == 0:
            build_ok = True
            break

    # Allow filesystem to flush the PDF if it exists
    for _ in range(10):
        if os.path.exists(pdf_path):
            build_ok = True
            break
        time.sleep(0.05)

    if not build_ok:
        print(f"Failed: {tex_file}")
        print("Log tail:\n" + "\n".join(last_stdout.splitlines()[-15:]))
        return False

    print(f"Generated: {os.path.basename(pdf_path)}\n")
    return True


def safe_filename(name):
    """Return a filesystem-safe basename for a given string."""
    return re.sub(r"[^\w\s-]", "", str(name or "Unknown")).strip().replace(
        " ", "_"
    )


def latex_escape_text(s):
    """
    Escape common LaTeX specials in plain text fields (e.g., Name).
    Do NOT use for math content.
    """
    if s is None:
        return "Unknown"
    repl = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(repl.get(ch, ch) for ch in str(s))


def clean_directory(directory, extensions):
    """Remove files by extension; prune empty subfolders."""
    for ext in extensions:
        for f in glob.glob(os.path.join(directory, f"**/*{ext}"),
                           recursive=True):
            try:
                os.remove(f)
            except Exception as e:
                print(f"Skipped {f}: {e}")

    for root, dirs, _ in os.walk(directory, topdown=False):
        for d in dirs:
            p = os.path.join(root, d)
            try:
                if not os.listdir(p):
                    shutil.rmtree(p, ignore_errors=True)
            except Exception:
                pass


# ============================================================
# MATH UTILITIES
# ============================================================



# ============================================================
# QUESTION BANK
# (Return LaTeX snippets ready to drop into @Qk@ placeholders)
# ============================================================
def Q1_Formula_Proof(n: int) -> str:
    arr = [
        r"We know that the Laplace transform of a function $f(t)$ is defined by \[ \mathcal{L}\{f(t)\}=\int_{0}^{\infty} e^{-st} f(t)\,dt\] Using this definition, show that \[ \mathcal{L}\{\sin(at)\}=\frac{a}{s^{2}+a^2}. \]",
        r"We know that the Laplace transform of a function $f(t)$ is defined by \[ \mathcal{L}\{f(t)\}=\int_{0}^{\infty} e^{-st} f(t)\,dt\] Using this definition, show that \[ \mathcal{L}\{\cos(at)\}=\frac{s}{s^{2}+a^2}. \]",
        r"We know that the Laplace transform of a function $f(t)$ is defined by \[ \mathcal{L}\{f(t)\}=\int_{0}^{\infty} e^{-st} f(t)\,dt\] Using this definition, show that \[ \mathcal{L}\{\sinh(at)\}=\frac{a}{s^{2}-a^2}. \]",
        r"We know that the Laplace transform of a function $f(t)$ is defined by \[ \mathcal{L}\{f(t)\}=\int_{0}^{\infty} e^{-st} f(t)\,dt\] Using this definition, show that \[ \mathcal{L}\{\cosh(at)\}=\frac{s}{s^{2}-a^2}. \]",
    ]
    s = arr[n - 1]
    return s


def Q2_Piecewise(n: int, a: int, b: int) -> str:
    arr = [
        r"Using definition, find the Laplace transform of the piecewise function \[ f(t)= \begin{cases} @a@ \sin @b@t, & 0\le t<\pi,\\[4pt] 0, & t\ge \pi. \end{cases} \]",
        r"Using definition, find the Laplace transform of the piecewise function \[ f(t)= \begin{cases} 0, & 0\le t<\pi,\\[4pt] @a@ \cos @b@t, & t\ge \pi. \end{cases} \]",
    ]
    s = arr[n - 1]
    s = s.replace("@a@", str(a)).replace("@b@", str(b))
    return s


def Q3_First_Translation(n: int, a: int, b: int, c: int) -> str:
    arr = [
        r"Find the Laplace transform of the piecewise function \[ f(t)= t e^{-@a@t} \sin(@b@t) \sin(@c@t) \]",
        r"Find the Laplace transform of the piecewise function \[ f(t)= t e^{@a@t} \cos(@b@t) \cos(@c@t) \]",
        r"Find the Laplace transform of the piecewise function \[ f(t)= t e^{-@a@t} \sin(@b@t) \cos(@c@t) \]",
        r"Find the Laplace transform of the piecewise function \[ f(t)= t e^{@a@t} \cos(@b@t) \sin(@c@t) \]",
    ]
    s = arr[n - 1]
    s = s = s.replace("@a@", str(a)).replace("@b@", str(b)).replace("@c@", str(c))
    return s


def Q4_Inverse_Translation() -> str:
    return r"Solve the Inverse Laplace problem\[\mathcal{L}^{-1}\left\{\frac{s}{s^{2}+2s-3}\right\}\]"


def Q5_Inverse_Partial(n: int) -> str:
    arr = [
        r"Solve the Inverse Laplace problem\[\mathcal{L}^{-1}\left\{\frac{2s-4}{(s^2+s)(s^2+1)}\right\}\]",
        r"Solve the Inverse Laplace problem\[\mathcal{L}^{-1}\left\{\frac{6s+3}{s^{4}+5s^{2}+4}\right\}\]",
    ]
    s = arr[n - 1]
    return s


def Q6_ODE_First_Order(a: int, b: int) -> str:    
    s = r"Use the Laplace transform to solve the given differential equation \[ y' + y = e^{-@a@t}\cos(@b@t),\qquad y(0)=0. \]"
    s = s.replace("@a@", str(a)).replace("@b@", str(b))
    return s


def Q7_ODE_Third_Order(n: int) -> str:
    arr = [
        r"Use the Laplace transform to solve the given differential equation \[ 2y''' + 3y'' - 3y' - 2y = e^{-t},\qquad y(0)=0,\; y'(0)=0,\; y''(0)=1. \]",
        r"Use the Laplace transform to solve the given differential equation \[ y''' + 2y'' - y' - 2y = \sin(3t),\qquad y(0)=0,\; y'(0)=0,\; y''(0)=1. \]",
    ]
    return arr[n - 1]


def Q8_ODE_Second_Order(n: int) -> str:
    arr = [
        r"Use the Laplace transform to solve the given differential equation \[ y'' + 9y = \cos 3t,\qquad y(0)=2,\; y'(0)=5. \]",
        r"Use the Laplace transform to solve the given differential equation \[y'' + y = \sin t,\qquad y(0)=1,\; y'(0)=-1.\]",
    ]
    s = arr[n - 1]
    return s


def Q9_ODE_Second_t_trig(n: int) -> str:
    arr = [
        r"Use the Laplace transform to solve the given differential equation \[ y' + y = t\sin t,\qquad y(0)=0. \]",
        r"Use the Laplace transform to solve the given differential equation \[ y' - y = t e^{t}\sin t,\qquad y(0)=0. \]",
    ]
    s = arr[n - 1]
    return s


def Q10_ODE_System(n: int) -> str:
    arr = [
        r"Use the Laplace transform to solve the given system of differential equations \[\begin{aligned}\frac{dx}{dt} &= x - 2y,\\\frac{dy}{dt} &= 5x - y,\end{aligned}\qquad x(0) = -1,\; y(0) = 2.\]",
        r"Use the Laplace transform to solve the given system of differential equations \[ \begin{aligned} \frac{dx}{dt} &= 2y + e^{t},\\ \frac{dy}{dt} &= 8x - t, \end{aligned} \qquad x(0) = 1,\; y(0) = 1. \]",
        r"Use the Laplace transform to solve the given system of differential equations \[ \begin{aligned} 2\frac{dx}{dt} + \frac{dy}{dt} - 2x &= 1,\\ \frac{dx}{dt} + \frac{dy}{dt} - 3x - 3y &= 2, \end{aligned} \qquad x(0)=0,\; y(0)=0. \]",
    ]
    s = arr[n - 1]
    return s




# ============================================================
# PREP
# ============================================================
start_time = time.time()
print(
    "LaTeX PDF Batch Generator | Started at "
    f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
)

output_dir = ASSESSMENT_TYPE
os.makedirs(output_dir, exist_ok=True)
print(f"Output directory: {output_dir}")

# Copy logo (some templates include it from working dir)
if os.path.exists(LOGO_FILE):
    shutil.copy(LOGO_FILE, output_dir)
    print(f"Copied logo → {output_dir}/")
else:
    print(f"Logo not found: '{LOGO_FILE}' (skipped)")

# Load workbook/template
try:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet = workbook[SHEET_NAME]
except Exception as e:
    raise SystemExit(f"Error loading workbook/sheet: {e}")

try:
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f_template:
        template = f_template.read().replace("\r\n", "\n")
except FileNotFoundError:
    raise SystemExit(f"Template not found: {TEMPLATE_PATH}")

print("\nStarting PDF generation...\n")


# ============================================================
# MAIN LOOP
# ============================================================
count_success, count_fail = 0, 0
row = START_ROW

while True:
    ID = sheet.cell(row=row, column=1).value
    Name = sheet.cell(row=row, column=2).value

    if not ID:
        print(
            f"\nCompleted processing. {count_success} succeeded, "
            f"{count_fail} failed."
        )
        break

    safe_name = safe_filename(Name)
    latex_name = latex_escape_text(Name)

    print(f"[{row - START_ROW + 1}] Processing: {Name}")

    Q1 = Q1_Formula_Proof(
        generate_integer(ID, "Q1_n", 1, 4),
    )

    Q2 = Q2_Piecewise(
        generate_integer(ID, "Q2_n", 1, 2),
        generate_integer(ID, "Q2_a", 2, 5),
        generate_integer(ID, "Q2_b", 2, 5),
    )

    Q3 = Q3_First_Translation(
        generate_integer(ID, "Q3_n", 1, 4),
        generate_integer(ID, "Q3_a", 2, 4),
        generate_integer(ID, "Q3_b", 2, 5),
        generate_integer(ID, "Q3_a", 2, 5),
    )

    Q4 = Q4_Inverse_Translation()

    Q5 = Q5_Inverse_Partial(
        generate_integer(ID, "Q5_n", 1, 2)
    )

    Q6 = Q6_ODE_First_Order(
        generate_integer(ID, "Q6_a", 2, 5),
        generate_integer(ID, "Q6_b", 2, 5),
    )

    Q7 = Q7_ODE_Third_Order(
        generate_integer(ID, "Q7_n", 1, 2)
    )

    Q8 = Q8_ODE_Second_Order(
        generate_integer(ID, "Q8_n", 1, 2),
    )

    Q9 = Q9_ODE_Second_t_trig(
        generate_integer(ID, "Q9_n", 1, 2),
    )

    Q10 = Q10_ODE_System(
        generate_integer(ID, "Q10_n", 1, 3),
    )

    tex_content = replace_placeholders(
        template,
        Name=latex_name,
        ID=ID,
        Section=SECTION,
        CourseName=COURSE_NAME,
        CourseCode=COURSE_CODE,
        SemesterName=SEMESTER_NAME,
        AssessmentType=ASSESSMENT_TYPE,
        TotalPoints=TOTAL_POINTS,

        Q1=Q1, Q2=Q2, Q3=Q3, Q4=Q4, Q5=Q5,
        Q6=Q6, Q7=Q7, Q8=Q8, Q9=Q9, Q10=Q10,
    )

    tex_filename = f"{ID}_{safe_name}.tex"
    tex_path = os.path.join(output_dir, tex_filename)

    with open(tex_path, "w", encoding="utf-8") as f_out:
        f_out.write(tex_content + "\n")

    if compile_latex_to_pdf(tex_path):
        count_success += 1
    else:
        count_fail += 1

    row += 1


# ============================================================
# CLEANUP
# ============================================================
print("\nPerforming cleanup...")
# Move any nested PDFs into output_dir root (defensive)
for root, _, files in os.walk(output_dir):
    for file in files:
        if file.lower().endswith(".pdf"):
            src = os.path.join(root, file)
            dest = os.path.join(output_dir, file)
            if src != dest:
                try:
                    shutil.move(src, dest)
                except Exception:
                    pass

# Remove aux + sources + logo
clean_directory(
    output_dir,
    [".aux", ".log", ".out", ".toc", ".nav", ".snm", ".bcf", ".xml", ".tex"],
)

logo_path = os.path.join(output_dir, LOGO_FILE)
if os.path.exists(logo_path):
    try:
        os.remove(logo_path)
    except Exception:
        pass

elapsed = time.time() - start_time
print(
    f"\nAll done! {count_success} PDFs generated successfully, "
    f"{count_fail} failed."
)
print(f"Clean folder ready at: {output_dir}")
print(
    f"Finished at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} "
    f"(Elapsed {elapsed:.2f}s)\n"
)
