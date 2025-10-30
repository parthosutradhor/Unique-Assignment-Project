import openpyxl
import hashlib
import re
import subprocess
import os
import shutil
import glob
import time
from datetime import datetime
from sympy import symbols, cos, sin, pi, sqrt, simplify, Rational, latex, I


# ==========================================
# CONFIGURATION
# ==========================================
ASSESSMENT_TYPE = "Assignment - 01"
SEMESTER_NAME = "FALL 2025"
COURSE_CODE = "MAT215"
COURSE_NAME = r"Complex Variables \& Laplace Transform"
SECTION = "12"
TOTAL_POINTS="100"

TEMPLATE_PATH = "question_template.tex"
WORKBOOK_PATH = "course-attendee.xlsx"
SHEET_NAME = "Worksheet"
LOGO_FILE = "Brac_University_Logo.png"

# Number of rows to skip before starting data
START_ROW = 35


# ==========================================
# UTILITY FUNCTIONS
# ==========================================
def generate_integers_range(ID, variation, n, a, b):
    """
    Generate deterministic pseudo-random integers from ID and variation.
    Ensures reproducibility based on hash seed.
    """
    seed = f"{ID}{variation}"
    hash_value = hashlib.md5(seed.encode()).hexdigest()
    return [(int(hash_value[i:i + 2], 32) % (b - a + 1)) + a for i in range(0, n * 2, 2)]


def replace_template_placeholders(template_text, **kwargs):
    """Replace placeholders like @Name@, @ID@, etc., in the LaTeX template."""
    for key, value in kwargs.items():
        template_text = template_text.replace(f"@{key}@", str(value))
    return template_text


def compile_latex_to_pdf(tex_path):
    """
    Compile a .tex file into a PDF using pdflatex.
    Runs twice for stable references.
    Shows log snippet if failed.
    """
    tex_dir = os.path.dirname(tex_path) or "."
    tex_file = os.path.basename(tex_path)

    cmd = ["pdflatex", "-interaction=nonstopmode", tex_file]

    for attempt in range(2):
        result = subprocess.run(
            cmd,
            cwd=tex_dir,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        if result.returncode != 0 and attempt == 1:
            print(f"⚠️  pdflatex returned non-zero exit code on {tex_file}")
            print("\n".join(result.stdout.splitlines()[-15:]))
            break

    pdf_path = os.path.splitext(tex_path)[0] + ".pdf"

    # Wait up to 1 second for PDF write
    for _ in range(10):
        if os.path.exists(pdf_path):
            print(f"✅ PDF generated successfully at: {pdf_path}")
            return True
        time.sleep(0.1)

    print(f"❌ PDF generation failed for {tex_file}")
    return False


def safe_filename(name):
    """Convert name to filesystem-safe format."""
    return re.sub(r"[^\w\s-]", "", str(name)).strip().replace(" ", "_")


def delete_files_by_extension(directory, extensions):
    """Delete all files in directory (recursive) matching given extensions."""
    for ext in extensions:
        for path in glob.glob(os.path.join(directory, f"**/*{ext}"), recursive=True):
            try:
                os.remove(path)
            except Exception as e:
                print(f"⚠️ Could not delete {path}: {e}")


def delete_empty_dirs(directory):
    """Delete all empty subdirectories inside the directory."""
    for root, dirs, _ in os.walk(directory, topdown=False):
        for d in dirs:
            dir_path = os.path.join(root, d)
            try:
                shutil.rmtree(dir_path)
                print(f"🗑️ Deleted inner folder: {dir_path}")
            except Exception as e:
                print(f"⚠️ Could not remove folder {dir_path}: {e}")





# ==========================================
# PREPARATION
# ==========================================
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
print(f"📄 LaTeX PDF Batch Generator | Started at {timestamp}\n")

output_dir = ASSESSMENT_TYPE
os.makedirs(output_dir, exist_ok=True)
print(f"📁 Output directory ready: {output_dir}")

# Copy logo if available
if os.path.exists(LOGO_FILE):
    shutil.copy(LOGO_FILE, output_dir)
    print(f"🖼️ Copied logo → {output_dir}/")
else:
    print(f"⚠️ Logo not found: '{LOGO_FILE}' (skipped)")

# Load workbook and template
workbook = openpyxl.load_workbook(WORKBOOK_PATH)
sheet = workbook[SHEET_NAME]

with open(TEMPLATE_PATH, "r", encoding="utf-8") as f_template:
    template = f_template.read().replace("\r\n", "\n")

print("\n🚀 Starting PDF generation...\n")


# ==========================================
# MAIN LOOP
# ==========================================
count_success, count_fail = 0, 0
row = START_ROW


# ==========================================
# Question Bank
# ==========================================

def complex_in_latex(r_val, theta_index):

    # --- Angle mapping (unit circle positions) ---
    angle_map = {
        0: 0,  1: 30,  2: 45,  3: 60,
        4: 90, 5: 120, 6: 135, 7: 150,
        8: 180, 9: 210, 10: 225, 11: 240,
        12: 270, 13: 300, 14: 315, 15: 330
    }

    if theta_index not in angle_map:
        raise ValueError("Angle index must be between 0 and 15.")

    theta_deg = Rational(angle_map[theta_index])

    # --- Convert modulus ---
    try:
        r = Rational(r_val)
    except Exception:
        r = simplify(r_val)

    # --- Conversion to radians ---
    theta_rad = theta_deg * pi / 180

    # --- Symbolic components ---
    x = simplify(r * cos(theta_rad))
    y = simplify(r * sin(theta_rad))

    # --- Symbolic complex number ---
    z = simplify(x + y * I)

    # --- Generate LaTeX output ---
    latex_expr = latex(z)
    latex_expr = latex_expr.replace(r'\operatorname{i}', 'i').replace(r'\mathrm{i}', 'i')

    return f"{latex_expr}"


def get_graphing_question_equality(n: int, a: str, b: str) -> str:
    arr = [
        r'\left|\frac{z+ai}{z-ai}\right|=b',
        r'|z+a|+|z-a|=2a+b',
        r'|z+ai|+|z-ai|=2a+b',
        r'|z-a|-|z+a|=2a-b',
        r'|z-ai|-|z+ai|=2a-b'
    ]

    def evaluate_expr(expr: str, a: str, b: str) -> str:
        try:
            expr = expr.replace('2a', '2*a').replace('2b', '2*b')
            a_val = float(a)
            b_val = float(b)
            expr = expr.replace('a', str(a_val)).replace('b', str(b_val))
            val = eval(expr)
            return str(int(val)) if val.is_integer() else str(val)
        except Exception:
            return expr

    s = arr[n - 1]

    # --- handle ai first ---
    s = s.replace('ai', f'{a}i').replace('+ai', f'+{a}i').replace('-ai', f'-{a}i')

    # --- detect and evaluate 2a±b patterns ---
    matches = re.findall(r'=(?:2a[+-]b)', s)
    for m in matches:
        expr = m[1:]
        val = evaluate_expr(expr, a, b)
        s = s.replace(m, f'={val}')

    # --- replace all a and b ---
    s = s.replace('{a}', f'{{{a}}}')
    s = s.replace('{b}', f'{{{b}}}')
    s = re.sub(r'(?<![A-Za-z])a(?![A-Za-z])', a, s)
    s = re.sub(r'(?<![A-Za-z])b(?![A-Za-z])', b, s)

    return f"{s}"




def get_graphing_question_inequality(n: int, a: str, b: str) -> str:
    arr = [
        r'\left|\frac{z+ai}{z-ai}\right| < b',
        r'\left|\frac{z+ai}{z-ai}\right| > b',
        r'\left|\frac{z+ai}{z-ai}\right| \le b',
        r'\left|\frac{z+ai}{z-ai}\right| \ge b',
        r'|z+a|+|z-a| < 2a+b',
        r'|z+a|+|z-a| > 2a+b',
        r'|z+a|+|z-a| \le 2a+b',
        r'|z+a|+|z-a| \ge 2a+b',
        r'|z+ai|+|z-ai| < 2a+b',
        r'|z+ai|+|z-ai| > 2a+b',
        r'|z+ai|+|z-ai| \le 2a+b',
        r'|z+ai|+|z-ai| \ge 2a+b',
        r'|z-a|-|z+a| < 2a-b',
        r'|z-a|-|z+a| > 2a-b',
        r'|z-a|-|z+a| \le 2a-b',
        r'|z-a|-|z+a| \ge 2a-b',
        r'|z-ai|-|z+ai| < 2a-b',
        r'|z-ai|-|z+ai| > 2a-b',
        r'|z-ai|-|z+ai| \le 2a-b',
        r'|z-ai|-|z+ai| \ge 2a-b'
    ]

    def evaluate_expr(expr: str, a: str, b: str) -> str:
        """Evaluate expressions like 2a+b or 2a-b when a,b are numeric."""
        try:
            expr = expr.replace('2a', '2*a').replace('2b', '2*b')
            a_val = float(a)
            b_val = float(b)
            expr = expr.replace('a', str(a_val)).replace('b', str(b_val))
            val = eval(expr)
            return str(int(val)) if val.is_integer() else str(val)
        except Exception:
            return expr  # keep symbolic if not numeric

    s = arr[n - 1]

    # --- handle ai first ---
    s = s.replace('+ai', f'+{a}i').replace('-ai', f'-{a}i')

    # --- detect and evaluate 2a±b patterns (for <, >, \le, \ge) ---
    pattern = r'(\\le|\\ge|<|>)\s*(2a[+-]b)'
    matches = re.findall(pattern, s)
    for symbol, expr in matches:
        val = evaluate_expr(expr, a, b)
        s = s.replace(f"{symbol} {expr}", f"{symbol} {val}")

    # --- replace all remaining a and b ---
    s = re.sub(r'(?<![A-Za-z])a(?![A-Za-z])', a, s)
    s = re.sub(r'(?<![A-Za-z])b(?![A-Za-z])', b, s)

    return f"{s}"



def get_prove_trig_hyp(n: int) -> str:
    arr = [
    r"\sin^{-1} z = \frac{1}{i} \ln \left( iz + \sqrt{1 - z^2} \right),",
    r"\cos^{-1} z = \frac{1}{i} \ln \left( z + \sqrt{z^2 - 1} \right),",
    r"\tan^{-1} z = \frac{1}{2i} \ln \left( \frac{1 + iz}{1 - iz} \right),",
    r"\cosec^{-1} z = \frac{1}{i} \ln \left( \frac{i + \sqrt{z^2 - 1}}{z} \right),",
    r"\sec^{-1} z = \frac{1}{i} \ln \left( \frac{1 + \sqrt{1 - z^2}}{z} \right),",
    r"\cot^{-1} z = \frac{1}{2i} \ln \left( \frac{z + i}{z - i} \right),",
    
    r"\sinh^{-1} z = \ln \left( z + \sqrt{z^2 + 1} \right),",
    r"\cosh^{-1} z = \ln \left( z + \sqrt{z^2 - 1} \right),",
    r"\tanh^{-1} z = \frac{1}{2} \ln \left( \frac{1 + z}{1 - z} \right),",
    r"\cosech^{-1} z = \ln \left( \frac{1 + \sqrt{z^2 + 1}}{z} \right),",
    r"\sech^{-1} z = \ln \left( \frac{1 + \sqrt{1 - z^2}}{z} \right),",
    r"\coth^{-1} z = \frac{1}{2} \ln \left( \frac{z + 1}{z - 1} \right)."
    ]

    s=arr[n-1]
    return f"{s}"


def get_solve_trig_hyp(n: int, a: int, b: int) -> str:
    arr = [
        r"\sin^{-1} z = a+bi",  r"\sin^{-1} z = a-bi",
        r"\cos^{-1} z = a+bi",  r"\cos^{-1} z = a-bi",
        r"\tan^{-1} z = a+bi",  r"\tan^{-1} z = a-bi",
        r"\cosec^{-1} z = a+bi",  r"\cosec^{-1} z = a-bi",
        r"\sec^{-1} z = a+bi",  r"\sec^{-1} z = a-bi",
        r"\cot^{-1} z = a+bi",  r"\cot^{-1} z = a-bi",
        r"\sinh^{-1} z = a+bi", r"\sinh^{-1} z = a-bi",
        r"\cosh^{-1} z = a+bi", r"\cosh^{-1} z = a-bi",
        r"\tanh^{-1} z = a+bi", r"\tanh^{-1} z = a-bi",
        r"\cosech^{-1} z = a+bi", r"\cosech^{-1} z = a-bi",
        r"\sech^{-1} z = a+bi", r"\sech^{-1} z = a-bi",
        r"\coth^{-1} z = a+bi", r"\coth^{-1} z = a-bi"
    ]

    # Get selected LaTeX template
    expr = arr[n-1]

    # Replace 'a' and 'b' with actual numeric values
    expr = re.sub(r'\ba\b', str(a), expr)
    expr = expr.replace('b', str(b))

    return expr


def Q10_get_derivative(n: int, a: int, b: int, c: int) -> str:
    arr = [
        r"Using the definition show that $$f(z)=@a@z^2 + @b@z - @c@$$ is differentiable at all points. Also find the derivative.",
        r"Using the definition show that $$f(z)=@a@z\bar{z} - @b@z + @c@\bar{z}$$ is not differentiable at $z=0$."
    ]

    # Get selected LaTeX template
    expr = arr[n-1]

    # Replace with actual numeric values
    expr = expr.replace('@a@', str(a)).replace('@b@', str(b)).replace('@c@', str(c))

    return expr


def Q11_get_derivative(n: int, a: int, b: int, c: int, d: int) -> str:
    arr = [
        r"Using the definition, find the derivative of \[f(z) = \frac{@a@z-@b@}{@c@z+@d@i} \quad \text{at} \quad z = i\].",
        r"Using the definition, find the derivative of \[f(z) = \frac{@a@}{@b@z + @c@} \quad \text{at} \quad z = z_0\].",
        r"Using the definition, find the derivative of \[f(z) = \frac{@a@}{z^2} \quad \text{at} \quad z = @b@+@c@i\]."
    ]

    # Get selected LaTeX template
    expr = arr[n-1]

    # Replace with actual numeric values
    expr = expr.replace('@a@', str(a)).replace('@b@', str(b)).replace('@c@', str(c)).replace('@d@', str(d))

    return expr







# ==========================================
# Placeholder Variables Replacement
# ==========================================




while True:
    ID = sheet.cell(row=row, column=1).value
    Name = sheet.cell(row=row, column=2).value

    if not ID:
        print(f"\n✅ Completed processing. {count_success} succeeded, {count_fail} failed.")
        break

    Name = Name or "Unknown"
    print(f"👤 Processing: {Name}")

    tex_content = replace_template_placeholders(
        template,
        Name=Name,
        ID=ID,
        Section=SECTION,
        Course_Name=COURSE_NAME,
        Course_Code=COURSE_CODE,
        Semester_Name=SEMESTER_NAME,
        Assesment_Type=ASSESSMENT_TYPE,
        Total_Points=TOTAL_POINTS,
        
        Q1_n=generate_integers_range(ID, "Q1_n", 1, 5, 7)[0],
        Q1_z=complex_in_latex((generate_integers_range(ID, "Q1_r", 1, 2, 3)[0])**(generate_integers_range(ID, "Q1_n", 1, 5, 7)[0]), generate_integers_range(ID, "Q1_arg", 1, 0, 15)[0]),

        graph_equation=get_graphing_question_equality(generate_integers_range(ID, "Q2_n", 1, 1, 5)[0], str(generate_integers_range(ID, "Q2_a", 1, 4, 9)[0],), str(generate_integers_range(ID, "Q2_b", 1, 1, 7)[0],)),

        graph_inequality=get_graphing_question_inequality(generate_integers_range(ID, "Q3_n", 1, 1, 20)[0], str(generate_integers_range(ID, "Q3_a", 1, 4, 9)[0],), str(generate_integers_range(ID, "Q3_b", 1, 1, 7)[0],)),

        Q4_a=generate_integers_range(ID, "Q4_a", 1, 2, 9)[0],

        Q4_z=complex_in_latex(generate_integers_range(ID, "Q4_r", 1, 2, 9)[0], generate_integers_range(ID, "Q4_arg", 1, 0, 15)[0]),

        Q5_expression=get_prove_trig_hyp(generate_integers_range(ID, "Q5_n", 1, 1, 12)[0]),

        Q6_expression=get_solve_trig_hyp(generate_integers_range(ID, "Q6_n", 1, 1, 24)[0], generate_integers_range(ID, "Q6_a", 1, 2, 9)[0], generate_integers_range(ID, "Q6_b", 1, 2, 9)[0]),

        Q10=Q10_get_derivative(generate_integers_range(ID, "Q10_n", 1, 1, 2)[0], generate_integers_range(ID, "Q10_a", 1, 2, 9)[0], generate_integers_range(ID, "Q10_b", 1, 2, 9)[0], generate_integers_range(ID, "Q10_c", 1, 2, 9)[0]),

        Q11=Q11_get_derivative(generate_integers_range(ID, "Q11_n", 1, 1, 3)[0], generate_integers_range(ID, "Q11_a", 1, 2, 9)[0], generate_integers_range(ID, "Q11_b", 1, 2, 9)[0], generate_integers_range(ID, "Q11_c", 1, 2, 9)[0], generate_integers_range(ID, "Q11_d", 1, 2, 9)[0])
    )





    safe_name = safe_filename(Name)
    tex_filename = f"{ID}_{safe_name}.tex"
    tex_path = os.path.join(output_dir, tex_filename)

    # Write LaTeX file
    with open(tex_path, "w", encoding="utf-8") as f_out:
        f_out.write(tex_content + "\n")

    if compile_latex_to_pdf(tex_path):
        count_success += 1
    else:
        count_fail += 1

    row += 1


# ==========================================
# CLEANUP
# ==========================================
print("\n🧹 Performing cleanup...")

# Move all PDFs up one level (if nested)
for root, _, files in os.walk(output_dir):
    for file in files:
        if file.lower().endswith(".pdf"):
            src = os.path.join(root, file)
            dest = os.path.join(output_dir, file)
            if src != dest:
                shutil.move(src, dest)

# Remove auxiliary LaTeX files and logo
delete_files_by_extension(output_dir, [".tex", ".aux", ".log", ".out"])
if os.path.exists(os.path.join(output_dir, LOGO_FILE)):
    os.remove(os.path.join(output_dir, LOGO_FILE))

# Remove subfolders
delete_empty_dirs(output_dir)

print(f"\n✅ All done! {count_success} PDFs generated successfully.")
print(f"📂 Clean folder ready at: {output_dir}")
print(f"🕓 Finished at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
