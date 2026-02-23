#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Author: Partho Sutra Dhor
Email: parthosutradhor@gmail.com, partho.dhor@bracu.ac.bd
Created: October 2025
Description: Automated LaTeX question booklet generator
"""

import glob
import hashlib
import os
import re
import shutil
import subprocess
import time
from datetime import datetime

import openpyxl
from sympy import I, Rational, cos, latex, pi, simplify, sin

# ============================================================
# CONFIGURATION
# ============================================================
ASSESSMENT_TYPE = "Assignment - 01"
SEMESTER_NAME = "Fall 2025"
COURSE_CODE = "MAT 215"
COURSE_NAME = r"Complex Variables and Laplace Transformations"
SECTION = "12"
TOTAL_POINTS = "150"

TEMPLATE_PATH = "assignment_01_template.tex"
WORKBOOK_PATH = "course-attendee.xlsx"
SHEET_NAME = "Worksheet"
LOGO_FILE = "Brac_University_Logo.png"

START_ROW = 2  # Data starts here (1-indexed, header on row 1)


# ============================================================
# GENERIC UTILITIES
# ============================================================
def generate_integers_range(ID, variation, n, a, b):
    """
    Deterministic pseudo-random integers based on (ID, variation).
    Uses MD5 -> hex pairs -> base-32 -> mod range mapping.
    """
    seed = f"{ID}{variation}"
    h = hashlib.md5(seed.encode()).hexdigest()
    return [
        (int(h[i:i + 2], 32) % (b - a + 1)) + a
        for i in range(0, n * 2, 2)
    ]


def replace_placeholders(template_text, **kwargs):
    """Replace @Key@ placeholders in a LaTeX template."""
    for key, val in kwargs.items():
        template_text = template_text.replace(f"@{key}@", str(val))
    return template_text


def compile_latex_to_pdf(tex_path):
    """
    Compile .tex -> .pdf using pdflatex twice.
    Prints a short log tail if compilation fails.
    """
    tex_dir = os.path.dirname(tex_path) or "."
    tex_file = os.path.basename(tex_path)
    pdf_path = os.path.splitext(tex_path)[0] + ".pdf"

    cmd = ["pdflatex", "-interaction=nonstopmode", tex_file]

    build_ok = False
    last_stdout = ""
    for _ in range(2):
        result = subprocess.run(
            cmd, cwd=tex_dir, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, text=True
        )
        last_stdout = result.stdout
        if result.returncode == 0:
            build_ok = True
            break

    # Allow filesystem to flush the PDF if it exists
    for _ in range(10):
        if os.path.exists(pdf_path):
            build_ok = True
            break
        time.sleep(0.05)

    if not build_ok:
        print(f"Failed: {tex_file}")
        print("Log tail:\n" + "\n".join(last_stdout.splitlines()[-15:]))
        return False

    print(f"Generated: {os.path.basename(pdf_path)}\n")
    return True


def safe_filename(name):
    """Return a filesystem-safe basename for a given string."""
    return re.sub(r"[^\w\s-]", "", str(name or "Unknown")).strip().replace(
        " ", "_"
    )


def latex_escape_text(s):
    """
    Escape common LaTeX specials in plain text fields (e.g., Name).
    Do NOT use for math content.
    """
    if s is None:
        return "Unknown"
    repl = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(repl.get(ch, ch) for ch in str(s))


def clean_directory(directory, extensions):
    """Remove files by extension; prune empty subfolders."""
    for ext in extensions:
        for f in glob.glob(os.path.join(directory, f"**/*{ext}"),
                           recursive=True):
            try:
                os.remove(f)
            except Exception as e:
                print(f"Skipped {f}: {e}")

    for root, dirs, _ in os.walk(directory, topdown=False):
        for d in dirs:
            p = os.path.join(root, d)
            try:
                if not os.listdir(p):
                    shutil.rmtree(p, ignore_errors=True)
            except Exception:
                pass


# ============================================================
# MATH UTILITIES
# ============================================================
def complex_in_latex(r_val, theta_index):
    """
    Return LaTeX for r·(cosθ + i·sinθ) at standard angles.
    θ index must be in [0, 15].
    """
    angle_map = {
        0: 0, 1: 30, 2: 45, 3: 60, 4: 90, 5: 120, 6: 135, 7: 150,
        8: 180, 9: 210, 10: 225, 11: 240, 12: 270, 13: 300, 14: 315, 15: 330,
    }
    if theta_index not in angle_map:
        raise ValueError("θ index must be between 0 and 15 (inclusive).")

    r = Rational(r_val)
    theta = Rational(angle_map[theta_index]) * pi / 180
    z = simplify(r * (cos(theta) + I * sin(theta)))
    return latex(z).replace(r"\operatorname{i}", "i").replace(r"\mathrm{i}", "i")


# ============================================================
# QUESTION BANK
# (Return LaTeX snippets ready to drop into @Qk@ placeholders)
# ============================================================
def Q1_get_nth_root(n: int, r_val: int, theta_index: int) -> str:
    rhs = complex_in_latex(r_val ** n, theta_index)
    return (
        r"Find all possible values of $z$ satisfying "
        rf"$$z^{{{n}}} = {rhs}.$$ "
        r"Locate them on the complex plane. Show that they lie on a circle, and determine its radius. Also, find the angular distance between two adjacent roots."
    )


def Q2_get_graphing_question_equality(n: int, a: int, b: int) -> str:
    arr = [
        r"\left|\frac{z+ai}{z-ai}\right|=@b@",
        r"|z+a|+|z-a|=2a+b",
        r"|z+ai|+|z-ai|=2a+b",
        r"|z-a|-|z+a|=2a-b",
        r"|z-ai|-|z+ai|=2a-b",
    ]
    s = arr[n - 1]
    s = (
        s.replace("@b@", str(b))
        .replace("z+a", "z+" + str(a))
        .replace("z-a", "z-" + str(a))
        .replace("2a+b", str(2 * a + b))
        .replace("2a-b", str(2 * a - b))
    )
    return rf"Describe the equation $\displaystyle {s}$ graphically on the complex plane."


def Q3_get_graphing_question_inequality(n: int, a: int, b: int) -> str:
    arr = [
        r"\left|\frac{z+ai}{z-ai}\right| < @b@",
        r"\left|\frac{z+ai}{z-ai}\right| > @b@",
        r"\left|\frac{z+ai}{z-ai}\right| \le @b@",
        r"\left|\frac{z+ai}{z-ai}\right| \ge @b@",
        r"|z+a|+|z-a| < 2a+b",
        r"|z+a|+|z-a| > 2a+b",
        r"|z+a|+|z-a| \le 2a+b",
        r"|z+a|+|z-a| \ge 2a+b",
        r"|z+ai|+|z-ai| < 2a+b",
        r"|z+ai|+|z-ai| > 2a+b",
        r"|z+ai|+|z-ai| \le 2a+b",
        r"|z+ai|+|z-ai| \ge 2a+b",
        r"|z-a|-|z+a| < 2a-b",
        r"|z-a|-|z+a| > 2a-b",
        r"|z-a|-|z+a| \le 2a-b",
        r"|z-a|-|z+a| \ge 2a-b",
        r"|z-ai|-|z+ai| < 2a-b",
        r"|z-ai|-|z+ai| > 2a-b",
        r"|z-ai|-|z+ai| \le 2a-b",
        r"|z-ai|-|z+ai| \ge 2a-b",
    ]
    s = arr[n - 1]
    s = (
        s.replace("@b@", str(b))
        .replace("z+a", "z+" + str(a))
        .replace("z-a", "z-" + str(a))
        .replace("2a+b", str(2 * a + b))
        .replace("2a-b", str(2 * a - b))
    )
    return rf"Describe the region $\displaystyle {s}$ graphically on the complex plane."


def Q4_get_solve_trig(a: int, r_val: int, theta_index: int) -> str:
    target = complex_in_latex(r_val, theta_index)
    return (
        r"Solve the equation "
        rf"$$e^{{{a}z}}={target}$$ "
        r"for $z$ and express $z$ as $x+iy$ where $x,y\in\mathbb{R}$."
    )


def Q5_get_prove_trig_hyp(n: int) -> str:
    arr = [
        r"\sin^{-1} z = \frac{1}{i}\,\ln\!\big( iz + \sqrt{1 - z^2} \big)",
        r"\cos^{-1} z = \frac{1}{i}\,\ln\!\big( z + \sqrt{z^2 - 1} \big)",
        r"\tan^{-1} z = \frac{1}{2i}\,\ln\!\left( \frac{1 + iz}{1 - iz} \right)",
        r"\cosec^{-1} z = \frac{1}{i}\,\ln\!\left( \frac{i + \sqrt{z^2 - 1}}{z} \right)",
        r"\sec^{-1} z = \frac{1}{i}\,\ln\!\left( \frac{1 + \sqrt{1 - z^2}}{z} \right)",
        r"\cot^{-1} z = \frac{1}{2i}\,\ln\!\left( \frac{z + i}{z - i} \right)",
        r"\sinh^{-1} z = \ln\!\big( z + \sqrt{z^2 + 1} \big)",
        r"\cosh^{-1} z = \ln\!\big( z + \sqrt{z^2 - 1} \big)",
        r"\tanh^{-1} z = \frac{1}{2}\,\ln\!\left( \frac{1 + z}{1 - z} \right)",
        r"\cosech^{-1} z = \ln\!\left( \frac{1 + \sqrt{z^2 + 1}}{z} \right)",
        r"\sech^{-1} z = \ln\!\left( \frac{1 + \sqrt{1 - z^2}}{z} \right)",
        r"\coth^{-1} z = \frac{1}{2}\,\ln\!\left( \frac{z + 1}{z - 1} \right)",
    ]
    s = arr[n - 1]
    return fr"Prove that $${s}.$$"


def Q6_get_solve_trig_hyp(n: int, a: int, b: int) -> str:
    arr = [
        r"\sin z = a+bi",  r"\sin z = a-bi",
        r"\cos z = a+bi",  r"\cos z = a-bi",
        r"\tan z = a+bi",  r"\tan z = a-bi",
        r"\cosec z = a+bi",  r"\cosec z = a-bi",
        r"\sec z = a+bi",  r"\sec z = a-bi",
        r"\cot z = a+bi",  r"\cot z = a-bi",
        r"\sinh z = a+bi", r"\sinh z = a-bi",
        r"\cosh z = a+bi", r"\cosh z = a-bi",
        r"\tanh z = a+bi", r"\tanh z = a-bi",
        r"\cosech z = a+bi", r"\cosech z = a-bi",
        r"\sech z = a+bi", r"\sech z = a-bi",
        r"\coth z = a+bi", r"\coth z = a-bi",
    ]
    s = arr[n - 1]
    s = s.replace("a+b", f"{a}+{b}").replace("a-b", f"{a}-{b}")
    return rf"Solve for $z$ where \[{s}.\]"


def Q7_get_limit_not_exists(n: int) -> str:
    arr = [
        r"Using the definition of a limit, show that $\displaystyle \lim_{z \to 0} \frac{\operatorname{Re}(z^2)}{|z|^2}$ does not exist.",
        r"Using the definition of a limit, show that $\displaystyle \lim_{z \to 0} \frac{\operatorname{Im}(z^2)}{|z|^2}$ does not exist.",
    ]
    return arr[n - 1]


def Q8_get_limit_LHopital(n: int, a: int, b: int) -> str:
    arr = [
        r"Using L’Hôpital’s rule, evaluate $$ \lim_{z \to 0} \left( \frac{\sin z}{z} \right)^{\frac{@a@ \sin(z)}{z - \sin z}}.$$",
        r"Using L’Hôpital’s rule, evaluate $$ \lim_{z \to 0} \left( \frac{\tan z}{z} \right)^{\frac{@a@ \sin(z)}{z - \sin z}}.$$",
        r"Using L’Hôpital’s rule, evaluate $$ \lim_{z \to 0} \left( \cos z \right)^{\frac{@a@ \sin(z)}{z - \sin z}}.$$",
        r"Using L’Hôpital’s rule, evaluate $$ \lim_{z \to 0} \left( \sec z \right)^{\frac{@a@ \sin(z)}{z - \sin z}}.$$",
    ]
    s = arr[n - 1]
    s = s.replace("@a@", str(a)).replace("@b@", str(b))
    return s


def Q9_get_Continuity(a: int, b: int) -> str:
    return (
        r"Consider the function "
        rf"\[f(z) = \frac{{\tan {a}z}}{{{b}z}}.\]"
        r"Is \( f(z) \) continuous at \( z = 0 \)? If not, redefine \( f \) at \( z = 0 \) so that \( f(z) \) becomes continuous. Also, find all points of discontinuity of \(f(z)\)."
    )


def Q10_get_derivative(n: int, a: int, b: int, c: int, d: int) -> str:
    arr = [
        r"Using the definition, find the derivative of $ \displaystyle f(z) = \frac{@a@z-@b@}{@c@z+@d@i} \quad \text{at} \quad z = i$.",
        r"Using the definition, find the derivative of $ \displaystyle f(z) = \frac{@a@}{@b@z + @c@} \quad \text{at} \quad z = z_0$.",
        r"Using the definition, find the derivative of $ \displaystyle f(z) = \frac{@a@}{z^2} \quad \text{at} \quad z = @b@+@c@i$.",
    ]
    s = arr[n - 1]
    s = (
        s.replace("@a@", str(a))
        .replace("@b@", str(b))
        .replace("@c@", str(c))
        .replace("@d@", str(d))
    )
    return s


def Q11_get_derivative(n: int, a: int, b: int, c: int) -> str:
    arr = [
        r"Using the definition, show that $$f(z)=@a@z^3 + @b@z - @c@$$ is differentiable at all points. Also find the derivative.",
        r"Using the definition, show that $$f(z)=@a@z\bar{z} - @b@z + @c@\bar{z}$$ is not differentiable at any point.",
    ]
    s = arr[n - 1]
    s = s.replace("@a@", str(a)).replace("@b@", str(b)).replace("@c@", str(c))
    return s


def Q12_get_analytic(n: int, a: int, b: int, c: int, d: int) -> str:
    arr = [
        r"Consider the function \[ f(z) = @a@ \sin(@b@z) - @c@ \cosh(@d@z).\] Using the Cauchy–Riemann equations, determine whether the function is analytic.",
        r"Consider the function \[ f(z) = @a@ \sinh(@b@z) - @c@ \cos(@d@z).\] Using the Cauchy–Riemann equations, determine whether the function is analytic.",
    ]
    s = arr[n - 1]
    s = (
        s.replace("@a@", str(a))
        .replace("@b@", str(b))
        .replace("@c@", str(c))
        .replace("@d@", str(d))
    )
    return s


def Q13_get_analytic(n: int, a: int, b: int, c: int) -> str:
    arr = [
        r"Consider the function \[ f(z) = @a@|z|^2 + @b@z - @c@\bar{z}.\] Using the Cauchy–Riemann equations, determine whether the function is analytic.",
        r"Consider the function \[ f(z) = @a@ze^{-@b@z}.\] Using the Cauchy–Riemann equations, determine whether the function is analytic.",
    ]
    s = arr[n - 1]
    s = (
        s.replace("@a@", str(a))
        .replace("@b@", str(b))
        .replace("@c@", str(c))
    )
    return s


def Q14_get_harmonic(n: int, a: int, b: int, c: int, d: int, e: int, f: int) -> str:
    arr = [
        r"Show that the function \[ U(x,y) = @a@ e^{-@b@x}\cos(@b@y)\;-\; @c@ e^{@d@y}\sin(@d@x) \;+\; @3e@\,x^2y \;-\; @f@x^2 \;-\; @e@y^3 \;+\; @f@y^2 \] is harmonic. Find the harmonic conjugate \textbf{$V$} of \textbf{$U$} such that \textbf{$U+Vi$} becomes analytic.",

        r"Show that the function \[ V(x,y) = @a@ e^{-@b@x}\cos(@b@y)\;-\; @c@ e^{@d@y}\sin(@d@x) \;+\; @3e@\,x^2y \;-\; @f@x^2 \;-\; @e@y^3 \;+\; @f@y^2 \] is harmonic. Find the harmonic conjugate \textbf{$U$} of \textbf{$V$} such that \textbf{$U+Vi$} becomes analytic.",

        r"Show that the function \[ U(x,y) = @a@ \sin(@b@x)\cosh(@b@y) \;+\; @3c@\,x^2y \;-\; @d@x^2 \;-\; @c@y^3 \;+\; @d@y^2 \] is harmonic. Find the harmonic conjugate \textbf{$V$} of \textbf{$U$} such that \textbf{$U+Vi$} becomes analytic.",

        r"Show that the function \[ V(x,y) = @a@ \sin(@b@x)\cosh(@b@y) \;+\; @3c@\,x^2y \;-\; @d@x^2 \;-\; @c@y^3 \;+\; @d@y^2 \] is harmonic. Find the harmonic conjugate \textbf{$U$} of \textbf{$V$} such that \textbf{$U+Vi$} becomes analytic.",
    ]
    s = arr[n - 1]
    s = (
        s.replace("@a@", str(a))
        .replace("@b@", str(b))
        .replace("@c@", str(c))
        .replace("@d@", str(d))
        .replace("@e@", str(e))
        .replace("@f@", str(f))
        .replace("@3c@", str(3 * c))
        .replace("@3e@", str(3 * e))
    )
    return s


def Q15_get_harmonic(n: int, a: int, b: int) -> str:
    arr = [
        r"Show that the function \[ U(x,y) = @a@\, x e^{-@b@x}\cos(@b@y) \;+\; @a@\, y e^{-@b@x}\sin(@b@y) \] is harmonic. Find the harmonic conjugate \textbf{$V$} of \textbf{$U$} such that \textbf{$U+Vi$} becomes analytic.",

        r"Show that the function \[ V(x,y) = @a@\, x e^{-@b@x}\cos(@b@y) \;+\; @a@\, y e^{-@b@x}\sin(@b@y) \] is harmonic. Find the harmonic conjugate \textbf{$U$} of \textbf{$V$} such that \textbf{$U+Vi$} becomes analytic."
    ]
    return arr[n - 1].replace("@a@", str(a)).replace("@b@", str(b))


# ============================================================
# PREP
# ============================================================
start_time = time.time()
print(
    "LaTeX PDF Batch Generator | Started at "
    f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
)

output_dir = ASSESSMENT_TYPE
os.makedirs(output_dir, exist_ok=True)
print(f"Output directory: {output_dir}")

# Copy logo (some templates include it from working dir)
if os.path.exists(LOGO_FILE):
    shutil.copy(LOGO_FILE, output_dir)
    print(f"Copied logo → {output_dir}/")
else:
    print(f"Logo not found: '{LOGO_FILE}' (skipped)")

# Load workbook/template
try:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet = workbook[SHEET_NAME]
except Exception as e:
    raise SystemExit(f"Error loading workbook/sheet: {e}")

try:
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f_template:
        template = f_template.read().replace("\r\n", "\n")
except FileNotFoundError:
    raise SystemExit(f"Template not found: {TEMPLATE_PATH}")

print("\nStarting PDF generation...\n")


# ============================================================
# MAIN LOOP
# ============================================================
count_success, count_fail = 0, 0
row = START_ROW

while True:
    ID = sheet.cell(row=row, column=1).value
    Name = sheet.cell(row=row, column=2).value

    if not ID:
        print(
            f"\nCompleted processing. {count_success} succeeded, "
            f"{count_fail} failed."
        )
        break

    safe_name = safe_filename(Name)
    latex_name = latex_escape_text(Name)

    print(f"[{row - START_ROW + 1}] Processing: {Name}")

    Q1 = Q1_get_nth_root(
        generate_integers_range(ID, "Q1_n", 1, 5, 7)[0],
        generate_integers_range(ID, "Q1_r", 1, 2, 3)[0],
        generate_integers_range(ID, "Q1_arg", 1, 0, 15)[0],
    )

    Q2 = Q2_get_graphing_question_equality(
        generate_integers_range(ID, "Q2_n", 1, 1, 5)[0],
        generate_integers_range(ID, "Q2_a", 1, 4, 9)[0],
        generate_integers_range(ID, "Q2_b", 1, 1, 7)[0],
    )

    Q3 = Q3_get_graphing_question_inequality(
        generate_integers_range(ID, "Q3_n", 1, 1, 20)[0],
        generate_integers_range(ID, "Q3_a", 1, 4, 9)[0],
        generate_integers_range(ID, "Q3_b", 1, 1, 7)[0],
    )

    Q4 = Q4_get_solve_trig(
        generate_integers_range(ID, "Q4_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q4_r", 1, 2, 9)[0],
        generate_integers_range(ID, "Q4_arg", 1, 0, 15)[0],
    )

    Q5 = Q5_get_prove_trig_hyp(
        generate_integers_range(ID, "Q5_n", 1, 1, 12)[0]
    )

    Q6 = Q6_get_solve_trig_hyp(
        generate_integers_range(ID, "Q6_n", 1, 1, 24)[0],
        generate_integers_range(ID, "Q6_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q6_b", 1, 2, 9)[0],
    )

    Q7 = Q7_get_limit_not_exists(
        generate_integers_range(ID, "Q7_n", 1, 1, 2)[0]
    )

    Q8 = Q8_get_limit_LHopital(
        generate_integers_range(ID, "Q8_n", 1, 1, 2)[0],
        generate_integers_range(ID, "Q8_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q8_b", 1, 2, 9)[0],
    )

    Q9 = Q9_get_Continuity(
        generate_integers_range(ID, "Q9_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q9_b", 1, 2, 9)[0],
    )

    Q10 = Q10_get_derivative(
        generate_integers_range(ID, "Q10_n", 1, 1, 3)[0],
        generate_integers_range(ID, "Q10_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q10_b", 1, 2, 9)[0],
        generate_integers_range(ID, "Q10_c", 1, 2, 9)[0],
        generate_integers_range(ID, "Q10_d", 1, 2, 9)[0],
    )

    Q11 = Q11_get_derivative(
        generate_integers_range(ID, "Q11_n", 1, 1, 2)[0],
        generate_integers_range(ID, "Q11_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q11_b", 1, 2, 9)[0],
        generate_integers_range(ID, "Q11_c", 1, 2, 9)[0],
    )

    Q12 = Q12_get_analytic(
        generate_integers_range(ID, "Q12_n", 1, 1, 2)[0],
        generate_integers_range(ID, "Q12_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q12_b", 1, 2, 9)[0],
        generate_integers_range(ID, "Q12_c", 1, 2, 9)[0],
        generate_integers_range(ID, "Q12_d", 1, 2, 9)[0],
    )

    Q13 = Q13_get_analytic(
        generate_integers_range(ID, "Q13_n", 1, 1, 2)[0],
        generate_integers_range(ID, "Q13_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q13_b", 1, 2, 9)[0],
        generate_integers_range(ID, "Q13_c", 1, 2, 9)[0],
    )

    Q14 = Q14_get_harmonic(
        generate_integers_range(ID, "Q14_n", 1, 1, 4)[0],
        generate_integers_range(ID, "Q14_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q14_b", 1, 2, 9)[0],
        generate_integers_range(ID, "Q14_c", 1, 2, 9)[0],
        generate_integers_range(ID, "Q14_d", 1, 2, 9)[0],
        generate_integers_range(ID, "Q14_e", 1, 2, 9)[0],
        generate_integers_range(ID, "Q14_f", 1, 2, 9)[0],
    )

    Q15 = Q15_get_harmonic(
        generate_integers_range(ID, "Q15_n", 1, 1, 2)[0],
        generate_integers_range(ID, "Q15_a", 1, 2, 9)[0],
        generate_integers_range(ID, "Q15_b", 1, 2, 9)[0],
    )

    tex_content = replace_placeholders(
        template,
        Name=latex_name,
        ID=ID,
        Section=SECTION,
        CourseName=COURSE_NAME,
        CourseCode=COURSE_CODE,
        SemesterName=SEMESTER_NAME,
        AssessmentType=ASSESSMENT_TYPE,
        TotalPoints=TOTAL_POINTS,

        Q1=Q1, Q2=Q2, Q3=Q3, Q4=Q4, Q5=Q5,
        Q6=Q6, Q7=Q7, Q8=Q8, Q9=Q9, Q10=Q10,
        Q11=Q11, Q12=Q12, Q13=Q13, Q14=Q14, Q15=Q15,
    )

    tex_filename = f"{ID}_{safe_name}.tex"
    tex_path = os.path.join(output_dir, tex_filename)

    with open(tex_path, "w", encoding="utf-8") as f_out:
        f_out.write(tex_content + "\n")

    if compile_latex_to_pdf(tex_path):
        count_success += 1
    else:
        count_fail += 1

    row += 1


# ============================================================
# CLEANUP
# ============================================================
print("\nPerforming cleanup...")
# Move any nested PDFs into output_dir root (defensive)
for root, _, files in os.walk(output_dir):
    for file in files:
        if file.lower().endswith(".pdf"):
            src = os.path.join(root, file)
            dest = os.path.join(output_dir, file)
            if src != dest:
                try:
                    shutil.move(src, dest)
                except Exception:
                    pass

# Remove aux + sources + logo
clean_directory(
    output_dir,
    [".aux", ".log", ".out", ".toc", ".nav", ".snm", ".bcf", ".xml", ".tex"],
)

logo_path = os.path.join(output_dir, LOGO_FILE)
if os.path.exists(logo_path):
    try:
        os.remove(logo_path)
    except Exception:
        pass

elapsed = time.time() - start_time
print(
    f"\nAll done! {count_success} PDFs generated successfully, "
    f"{count_fail} failed."
)
print(f"Clean folder ready at: {output_dir}")
print(
    f"Finished at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} "
    f"(Elapsed {elapsed:.2f}s)\n"
)
